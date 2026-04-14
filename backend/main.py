from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
import json
import pandas as pd
from datetime import datetime, timedelta
import io
from functools import lru_cache
from pathlib import Path
import math
import os
import shutil
from threading import Lock, Thread
from tempfile import NamedTemporaryFile
from typing import Any
from uuid import uuid4
import zipfile

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import BaseModel

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_DIR = Path(__file__).resolve().parent.parent
HISTORY_FILE = BASE_DIR / "data" / "history.xlsx"
HISTORY_PENDING_FILE = BASE_DIR / "data" / "history.pending.jsonl"
ACTIVITY_FILE = BASE_DIR / "data" / "activity.jsonl"
RECENT_APPLICATION_WINDOW_DAYS = 14
ANALYSIS_SESSIONS: dict[str, dict] = {}
HISTORY_FILE_LOCK = Lock()
HISTORY_PENDING_LOCK = Lock()
ACTIVITY_LOCK = Lock()
HISTORY_FLUSH_LOCK = Lock()
HISTORY_THREAD_LOCK = Lock()
HISTORY_FLUSH_THREAD: Thread | None = None


class DecisionPayload(BaseModel):
    decision: str


class HistoryWorkbookError(Exception):
    pass


def build_history_warning():
    return (
        f"{HISTORY_FILE.name} is unreadable or incomplete. "
        "History checks are temporarily running with no prior records until the workbook is restored."
    )


def read_history_excel():
    try:
        return pd.read_excel(HISTORY_FILE, header=1)
    except (zipfile.BadZipFile, ValueError, OSError) as exc:
        raise HistoryWorkbookError(build_history_warning()) from exc


def save_workbook_atomic(workbook: Workbook, destination: Path):
    destination.parent.mkdir(parents=True, exist_ok=True)
    temp_path = None
    try:
        with NamedTemporaryFile(
            delete=False,
            dir=destination.parent,
            suffix=destination.suffix,
        ) as temp_file:
            temp_path = Path(temp_file.name)

        workbook.save(temp_path)
        os.replace(temp_path, destination)
    finally:
        if temp_path is not None and temp_path.exists():
            temp_path.unlink(missing_ok=True)


def quarantine_corrupt_history_file():
    if not HISTORY_FILE.exists():
        return None

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = HISTORY_FILE.with_name(f"{HISTORY_FILE.stem}.corrupt.{timestamp}{HISTORY_FILE.suffix}")
    os.replace(HISTORY_FILE, backup_path)
    return backup_path

def get_file_signature(path: Path):
    if not path.exists():
        return None

    file_stat = path.stat()
    return (file_stat.st_mtime_ns, file_stat.st_size)


def get_pending_history_paths_locked():
    if not HISTORY_PENDING_FILE.parent.exists():
        return []

    pending_paths = []
    processing_pattern = f"{HISTORY_PENDING_FILE.stem}.processing.*{HISTORY_PENDING_FILE.suffix}"
    pending_paths.extend(sorted(HISTORY_PENDING_FILE.parent.glob(processing_pattern)))

    if HISTORY_PENDING_FILE.exists():
        pending_paths.append(HISTORY_PENDING_FILE)

    return pending_paths


def get_pending_history_signature():
    with HISTORY_PENDING_LOCK:
        pending_paths = get_pending_history_paths_locked()
        if not pending_paths:
            return None

        return tuple(
            (path.name, path.stat().st_mtime_ns, path.stat().st_size)
            for path in pending_paths
            if path.exists()
        )


def read_pending_history_rows_from_file(path: Path):
    if not path.exists():
        return []

    rows = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            payload = line.strip()
            if not payload:
                continue

            try:
                row = json.loads(payload)
            except json.JSONDecodeError:
                continue

            if isinstance(row, dict):
                rows.append(row)

    return rows


def read_pending_history_rows():
    with HISTORY_PENDING_LOCK:
        rows = []
        for path in get_pending_history_paths_locked():
            rows.extend(read_pending_history_rows_from_file(path))
        return rows


def merge_history_rows(history_df: pd.DataFrame, pending_rows: list[dict]):
    if not pending_rows:
        return history_df

    pending_df = pd.DataFrame(pending_rows)
    if pending_df.empty:
        return history_df

    if history_df.empty:
        return pending_df

    merged_columns = list(dict.fromkeys([*history_df.columns.tolist(), *pending_df.columns.tolist()]))
    return pd.concat(
        [
            history_df.reindex(columns=merged_columns),
            pending_df.reindex(columns=merged_columns),
        ],
        ignore_index=True,
    )


@lru_cache(maxsize=4)
def _load_history_cached(
    history_signature: tuple[int, int] | None,
    pending_signature: tuple[int, int] | None,
):
    history_df = read_history_excel() if history_signature is not None else pd.DataFrame()
    return merge_history_rows(history_df, read_pending_history_rows())

def load_history():
    history_signature = get_file_signature(HISTORY_FILE)
    pending_signature = get_pending_history_signature()
    if history_signature is None and pending_signature is None:
        return pd.DataFrame()

    try:
        return _load_history_cached(history_signature, pending_signature)
    except HistoryWorkbookError:
        return merge_history_rows(pd.DataFrame(), read_pending_history_rows())


def load_history_with_status():
    history_signature = get_file_signature(HISTORY_FILE)
    pending_signature = get_pending_history_signature()
    if history_signature is None and pending_signature is None:
        return pd.DataFrame(), None

    try:
        return _load_history_cached(history_signature, pending_signature), None
    except HistoryWorkbookError as exc:
        return merge_history_rows(pd.DataFrame(), read_pending_history_rows()), str(exc)

def prepare_history_for_matching(history_df: pd.DataFrame):
    prepared = history_df.copy()

    if "BOOK_DATE" in prepared.columns:
        prepared["BOOK_DATE"] = pd.to_datetime(prepared["BOOK_DATE"], errors="coerce")
    else:
        prepared["BOOK_DATE"] = pd.NaT

    if "EC_NUMBER" in prepared.columns:
        prepared["_ec_key"] = prepared["EC_NUMBER"].map(normalize_identifier)
    else:
        prepared["_ec_key"] = None

    if "CUSTOMER_NO" in prepared.columns:
        prepared["_customer_key"] = prepared["CUSTOMER_NO"].map(normalize_identifier)
    else:
        prepared["_customer_key"] = None

    lookup: dict[str, list[int]] = {}
    for key_column in ("_ec_key", "_customer_key"):
        valid_entries = prepared[key_column].dropna()
        for row_index, key in valid_entries.items():
            lookup.setdefault(key, []).append(row_index)

    return prepared, lookup


@lru_cache(maxsize=4)
def _load_history_for_matching_cached(
    history_signature: tuple[int, int] | None,
    pending_signature: tuple[int, int] | None,
):
    history_df = read_history_excel() if history_signature is not None else pd.DataFrame()
    return prepare_history_for_matching(merge_history_rows(history_df, read_pending_history_rows()))

def load_history_for_matching():
    history_signature = get_file_signature(HISTORY_FILE)
    pending_signature = get_pending_history_signature()
    if history_signature is None and pending_signature is None:
        return pd.DataFrame(columns=["BOOK_DATE"]), {}

    try:
        return _load_history_for_matching_cached(history_signature, pending_signature)
    except HistoryWorkbookError:
        return prepare_history_for_matching(
            merge_history_rows(pd.DataFrame(columns=["BOOK_DATE"]), read_pending_history_rows())
        )


def load_history_for_matching_with_status():
    history_signature = get_file_signature(HISTORY_FILE)
    pending_signature = get_pending_history_signature()
    if history_signature is None and pending_signature is None:
        return pd.DataFrame(columns=["BOOK_DATE"]), {}, None

    try:
        history_df, history_lookup = _load_history_for_matching_cached(history_signature, pending_signature)
        return history_df, history_lookup, None
    except HistoryWorkbookError as exc:
        history_df, history_lookup = prepare_history_for_matching(
            merge_history_rows(pd.DataFrame(columns=["BOOK_DATE"]), read_pending_history_rows())
        )
        return history_df, history_lookup, str(exc)

def normalize_scalar(value):
    if hasattr(value, "item") and not isinstance(value, (str, bytes)):
        try:
            value = value.item()
        except (ValueError, TypeError):
            pass

    if pd.isna(value):
        return None

    if isinstance(value, str):
        value = value.strip()
        return value or None

    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None

    return value

def normalize_column_name(column_name):
    return "".join(character for character in str(column_name).lower() if character.isalnum())

def normalize_identifier(value):
    value = normalize_scalar(value)
    if value is None:
        return None

    if isinstance(value, float) and value.is_integer():
        value = int(value)

    normalized = str(value).strip().upper()
    if normalized.endswith(".0") and normalized[:-2].isdigit():
        normalized = normalized[:-2]

    normalized = "".join(normalized.split())
    return normalized or None

def normalize_name(name):
    """Normalize name for comparison by removing extra spaces and converting to uppercase."""
    if not name or pd.isna(name):
        return ""
    return " ".join(str(name).upper().split())

def names_are_similar(name1, name2, threshold=0.75):
    """Check if two names are similar (likely the same person with minor variations)."""
    n1 = normalize_name(name1)
    n2 = normalize_name(name2)

    if n1 == n2:
        return True

    if not n1 or not n2:
        return False

    # Split into words for better comparison
    words1 = set(n1.split())
    words2 = set(n2.split())

    # If one name is subset of another (e.g., middle name added/removed)
    if words1.issubset(words2) or words2.issubset(words1):
        return True

    # Check for significant overlap - at least 60% of shorter name's words
    common_words = words1.intersection(words2)
    shorter_words = min(len(words1), len(words2))
    if shorter_words > 0 and len(common_words) / shorter_words >= 0.6:
        return True

    # For character-level similarity, require substantial consecutive matches
    shorter = min(n1, n2, key=len)
    longer = max(n1, n2, key=len)

    if len(shorter) == 0 or len(longer) == 0:
        return False

    # Don't compare if length difference is too large
    if len(longer) - len(shorter) > max(2, len(shorter) * 0.2):  # Allow 20% length difference max
        return False

    # Find longest common substring
    def longest_common_substring(s1, s2):
        m, n = len(s1), len(s2)
        dp = [[0] * (n + 1) for _ in range(m + 1)]
        max_length = 0

        for i in range(1, m + 1):
            for j in range(1, n + 1):
                if s1[i - 1] == s2[j - 1]:
                    dp[i][j] = dp[i - 1][j - 1] + 1
                    max_length = max(max_length, dp[i][j])

        return max_length

    lcs_length = longest_common_substring(n1, n2)
    lcs_similarity = lcs_length / len(shorter)

    # Simple edit distance for small differences
    def edit_distance(s1, s2):
        m, n = len(s1), len(s2)
        dp = [[0] * (n + 1) for _ in range(m + 1)]

        for i in range(m + 1):
            dp[i][0] = i
        for j in range(n + 1):
            dp[0][j] = j

        for i in range(1, m + 1):
            for j in range(1, n + 1):
                cost = 0 if s1[i - 1] == s2[j - 1] else 1
                dp[i][j] = min(
                    dp[i - 1][j] + 1,      # deletion
                    dp[i][j - 1] + 1,      # insertion
                    dp[i - 1][j - 1] + cost  # substitution
                )

        return dp[m][n]

    edit_dist = edit_distance(n1, n2)
    edit_similarity = 1 - (edit_dist / max(len(n1), len(n2)))

    # Use the better of LCS or edit distance similarity
    best_similarity = max(lcs_similarity, edit_similarity)

    return best_similarity >= threshold

def detect_similar_names(names_list):
    """Check if any names in the list are similar (potential fraud)."""
    names_list = [n for n in names_list if n and not pd.isna(n)]
    if len(names_list) < 2:
        return False

    for i, name1 in enumerate(names_list):
        for j, name2 in enumerate(names_list):
            if i != j and names_are_similar(name1, name2):
                return True
    return False

def parse_date(value):
    value = normalize_scalar(value)
    if value is None:
        return None

    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None

    if isinstance(parsed, pd.Timestamp):
        return parsed.to_pydatetime()

    return parsed


def determine_initial_decision_status(row: pd.Series) -> str | None:
    raw_response = get_first_present(
        row,
        "RESPONSE",
        "STATUS",
        "DECISION",
        "RESULT",
        "APPLICATION_STATUS",
        "RESPONSE_STATUS",
        "BeneficiaryStatus",
    )
    if raw_response is None:
        return None

    response = str(raw_response).strip().lower()
    if not response:
        return None

    if any(token in response for token in ["accept", "approve", "yes", "true", "1", "ok", "passed", "completed", "in process"]):
        return "approved"
    if any(token in response for token in ["reject", "decline", "deny", "no", "false", "0", "failed", "denied"]):
        return "declined"
    if any(token in response for token in ["review", "hold", "pending"]):
        return "manual_review"

    return None


def to_jsonable_records(df: pd.DataFrame) -> list[dict]:
    if df.empty:
        return []

    visible_columns = [column for column in df.columns if not str(column).startswith("_")]
    cleaned = df[visible_columns].replace([math.inf, -math.inf], pd.NA).astype(object)
    cleaned = cleaned.where(pd.notna(cleaned), None)

    records = []
    for record in cleaned.to_dict(orient="records"):
        jsonable_record = {}
        for key, value in record.items():
            value = normalize_scalar(value)
            if isinstance(value, (datetime, pd.Timestamp)):
                value = value.isoformat()
            jsonable_record[key] = value
        records.append(jsonable_record)

    return records

def get_first_present(row: pd.Series, *candidate_names: str):
    normalized_columns = {
        normalize_column_name(column): column
        for column in row.index
    }

    for candidate in candidate_names:
        source_column = normalized_columns.get(normalize_column_name(candidate))
        if source_column is None:
            continue

        value = normalize_scalar(row.get(source_column))
        if value is not None:
            return value

    return None


def to_excel_cell(value):
    value = normalize_scalar(value)
    if isinstance(value, pd.Timestamp):
        value = value.to_pydatetime()
    return value


def to_history_journal_cell(value):
    value = to_excel_cell(value)
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def make_json_safe(value: Any):
    if isinstance(value, dict):
        return {
            str(key): make_json_safe(item)
            for key, item in value.items()
        }

    if isinstance(value, (list, tuple)):
        return [make_json_safe(item) for item in value]

    value = normalize_scalar(value)
    if isinstance(value, (datetime, pd.Timestamp)):
        return value.isoformat()

    return value


def row_to_storage(row: pd.Series, columns: list) -> dict:
    return {
        str(column): to_excel_cell(row.get(column))
        for column in columns
    }


def create_history_workbook(columns: list[str]):
    HISTORY_FILE.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(["NEW LOANS CAPTURED"])
    worksheet.append(columns)
    save_workbook_atomic(workbook, HISTORY_FILE)
    workbook.close()


def ensure_history_headers() -> list[str]:
    if not HISTORY_FILE.exists():
        create_history_workbook([])
        return []

    workbook = load_workbook(HISTORY_FILE)
    worksheet = workbook.active
    headers = [
        "" if header is None else str(header)
        for header in next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True), ())
    ]
    workbook.close()
    return headers


def clear_history_caches():
    _load_history_cached.cache_clear()
    _load_history_for_matching_cached.cache_clear()


def append_records_to_history_workbook(raw_rows: list[dict]):
    if not raw_rows:
        return 0

    with HISTORY_FILE_LOCK:
        if not HISTORY_FILE.exists():
            create_history_workbook(list(raw_rows[0].keys()))
        else:
            try:
                workbook = load_workbook(HISTORY_FILE)
                workbook.close()
            except (zipfile.BadZipFile, InvalidFileException, OSError, KeyError):
                quarantine_corrupt_history_file()
                create_history_workbook(list(raw_rows[0].keys()))

        workbook = load_workbook(HISTORY_FILE)
        try:
            worksheet = workbook.active
            history_headers = [
                "" if header is None else str(header)
                for header in next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True), ())
            ]

            if not history_headers:
                history_headers = list(raw_rows[0].keys())
                if worksheet.max_row < 2:
                    if worksheet.max_row == 0:
                        worksheet.append(["NEW LOANS CAPTURED"])
                    worksheet.append(history_headers)
                else:
                    for column_index, header in enumerate(history_headers, start=1):
                        worksheet.cell(row=2, column=column_index, value=header)

            normalized_headers = [normalize_column_name(header) for header in history_headers]
            for raw_row in raw_rows:
                normalized_uploaded = {
                    normalize_column_name(column): to_excel_cell(value)
                    for column, value in raw_row.items()
                }
                worksheet.append([
                    normalized_uploaded.get(header)
                    for header in normalized_headers
                ])

            save_workbook_atomic(workbook, HISTORY_FILE)
        finally:
            workbook.close()

    clear_history_caches()
    return len(raw_rows)


def claim_pending_history_file():
    with HISTORY_PENDING_LOCK:
        if not HISTORY_PENDING_FILE.exists():
            return None

        if HISTORY_PENDING_FILE.stat().st_size == 0:
            HISTORY_PENDING_FILE.unlink(missing_ok=True)
            return None

        processing_path = HISTORY_PENDING_FILE.with_name(
            f"{HISTORY_PENDING_FILE.stem}.processing.{uuid4().hex}{HISTORY_PENDING_FILE.suffix}"
        )
        os.replace(HISTORY_PENDING_FILE, processing_path)
        return processing_path


def merge_pending_history_file_back(processing_path: Path):
    if not processing_path.exists():
        return

    merged_path = HISTORY_PENDING_FILE.with_name(
        f"{HISTORY_PENDING_FILE.stem}.merge.{uuid4().hex}{HISTORY_PENDING_FILE.suffix}"
    )

    try:
        with HISTORY_PENDING_LOCK:
            with merged_path.open("wb") as destination:
                with processing_path.open("rb") as source:
                    shutil.copyfileobj(source, destination)

                if HISTORY_PENDING_FILE.exists():
                    with HISTORY_PENDING_FILE.open("rb") as current:
                        shutil.copyfileobj(current, destination)

            os.replace(merged_path, HISTORY_PENDING_FILE)
    finally:
        if merged_path.exists():
            merged_path.unlink(missing_ok=True)


def flush_pending_history_rows():
    total_flushed = 0
    with HISTORY_FLUSH_LOCK:
        while True:
            processing_path = claim_pending_history_file()
            if processing_path is None:
                return total_flushed

            try:
                pending_rows = read_pending_history_rows_from_file(processing_path)
                if pending_rows:
                    total_flushed += append_records_to_history_workbook(pending_rows)
            except Exception:
                merge_pending_history_file_back(processing_path)
                raise
            finally:
                processing_path.unlink(missing_ok=True)


def _background_flush_pending_history():
    global HISTORY_FLUSH_THREAD

    try:
        flush_pending_history_rows()
    finally:
        with HISTORY_THREAD_LOCK:
            HISTORY_FLUSH_THREAD = None


def schedule_history_flush():
    global HISTORY_FLUSH_THREAD

    with HISTORY_THREAD_LOCK:
        if HISTORY_FLUSH_THREAD is not None and HISTORY_FLUSH_THREAD.is_alive():
            return

        HISTORY_FLUSH_THREAD = Thread(
            target=_background_flush_pending_history,
            name="history-flush",
            daemon=True,
        )
        HISTORY_FLUSH_THREAD.start()


def append_record_to_history(raw_row: dict):
    serialized_row = {
        str(column): to_history_journal_cell(value)
        for column, value in raw_row.items()
    }

    HISTORY_PENDING_FILE.parent.mkdir(parents=True, exist_ok=True)
    with HISTORY_PENDING_LOCK:
        with HISTORY_PENDING_FILE.open("a", encoding="utf-8") as handle:
            handle.write(json.dumps(serialized_row, ensure_ascii=True))
            handle.write("\n")

    clear_history_caches()
    schedule_history_flush()


def build_download_dataframe(session: dict, decision_status: str) -> pd.DataFrame:
    selected_rows = [
        session["rows_by_id"][application_id]
        for application_id in session["record_order"]
        if session["records_by_id"][application_id]["decision_status"] == decision_status
    ]
    return pd.DataFrame(selected_rows, columns=session["columns"])


def build_decision_summary(session: dict) -> dict:
    records = session["records_by_id"].values()
    return {
        "approved_count": sum(record["decision_status"] == "approved" for record in records),
        "rejected_count": sum(record["decision_status"] == "declined" for record in records),
        "pending_count": sum(
            record["decision_status"] in {"pending", "manual_review"}
            for record in records
        ),
    }


def serialize_record(record: dict) -> dict:
    response_record = {
        key: value
        for key, value in record.items()
        if key not in {"raw_row"}
    }
    return response_record


def build_session_response(session: dict) -> dict:
    records = [
        serialize_record(session["records_by_id"][application_id])
        for application_id in session["record_order"]
    ]
    anomalies = [record for record in records if record["category"] == "anomaly"]
    clear_records = [record for record in records if record["category"] == "clear"]
    decision_summary = build_decision_summary(session)

    return {
        "session_id": session["session_id"],
        "analysis_mode": session["analysis_mode"],
        "file_name": session["file_name"],
        "total_processed": session["total_processed"],
        "actionable_records": len(records),
        "window_days": RECENT_APPLICATION_WINDOW_DAYS,
        **decision_summary,
        "history_warning": session.get("history_warning"),
        "anomalies": anomalies,
        "clear_records": clear_records,
    }


def get_session_or_404(session_id: str) -> dict:
    session = ANALYSIS_SESSIONS.get(session_id)
    if session is None:
        raise HTTPException(status_code=404, detail="Analysis session not found.")
    return session

def build_application_record(row: pd.Series, index: int):
    row_number = index + 2
    applicant_name = get_first_present(
        row,
        "CUSTOMER_NAME1",
        "CUSTOMER NAME1",
        "CUSTOMER NAME",
        "FULL NAME",
        "NAME",
        "APPLICANT NAME",
        "CLIENT NAME",
        "BORROWER NAME",
    )
    ec_number = get_first_present(
        row,
        "EC_NUMBER",
        "EC NUMBER",
        "EC NO",
        "EC",
        "ECONOMIC CENTER",
        "ECONOMIC CENTRE",
        "BRANCH CODE",
        "BRANCH",
    )
    customer_no = get_first_present(
        row,
        "CUSTOMER_NO",
        "CUSTOMER NO",
        "ID",
        "ID NUMBER",
        "CUSTOMER ID",
        "CLIENT ID",
        "ACCOUNT NUMBER",
        "ACCOUNT NO",
        "ACCOUNT",
    )
    amount = get_first_present(
        row,
        "AMOUNT_FINANCED",
        "AMOUNT FINANCED",
        "AMOUNT",
        "LOAN AMOUNT",
        "FINANCE AMOUNT",
        "CREDIT AMOUNT",
        "VALUE",
    )
    application_book_date = parse_date(
        get_first_present(
            row,
            "BOOK_DATE",
            "BOOK DATE",
            "APPLICATION DATE",
            "DATE",
            "LOAN DATE",
            "DISBURSEMENT DATE",
            "APPROVAL DATE",
        )
    )

    return {
        "application_id": f"row-{row_number}",
        "row": row_number,
        "applicant_name": applicant_name,
        "ec_number": ec_number,
        "customer_no": customer_no,
        "amount": amount,
        "application_book_date": application_book_date.isoformat() if application_book_date else None,
        "decision_status": "pending",
    }

def analyze_application(
    row: pd.Series,
    index: int,
    history_df: pd.DataFrame,
    history_lookup: dict[str, list[int]],
):
    application = build_application_record(row, index)

    identifier_keys = {
        normalize_identifier(application["ec_number"]),
        normalize_identifier(application["customer_no"]),
    }
    identifier_keys.discard(None)

    reference_date = parse_date(application["application_book_date"]) or datetime.now()
    review_window_start = reference_date - timedelta(days=RECENT_APPLICATION_WINDOW_DAYS)

    anomaly_reasons = []

    if not identifier_keys:
        anomaly_reasons.append("Missing EC number / customer ID. Unable to check recent history.")
        application.update({
            "category": "anomaly",
            "anomaly_reasons": anomaly_reasons,
            "reason": anomaly_reasons[0],
            "reference_date": reference_date.isoformat(),
            "history_match_count": 0,
            "recent_match_count": 0,
            "latest_book_date": None,
            "matched_records": [],
        })
        return application

    matched_indices = set()
    for key in identifier_keys:
        matched_indices.update(history_lookup.get(key, []))

    if matched_indices:
        matched_records = history_df.loc[sorted(matched_indices)].sort_values("BOOK_DATE", ascending=False)
    else:
        matched_records = history_df.iloc[0:0].copy()

    recent_matches = matched_records[
        matched_records["BOOK_DATE"].notna() &
        matched_records["BOOK_DATE"].between(review_window_start, reference_date)
    ].sort_values("BOOK_DATE", ascending=False)

    historical_matches = matched_records[
        matched_records["BOOK_DATE"].notna() &
        (matched_records["BOOK_DATE"] < review_window_start)
    ].sort_values("BOOK_DATE", ascending=False)

    latest_book_date = matched_records["BOOK_DATE"].dropna().max() if not matched_records.empty else None
    latest_book_date_iso = latest_book_date.isoformat() if pd.notna(latest_book_date) else None

    if not recent_matches.empty:
        anomaly_reasons.append(
            f"Previous loan found within the last {RECENT_APPLICATION_WINDOW_DAYS} days."
        )

    # Check for identity fraud: same EC number or account number used by different or similar names
    if not matched_records.empty:
        current_name = normalize_name(application["applicant_name"])
        app_ec_key = normalize_identifier(application["ec_number"])
        app_customer_key = normalize_identifier(application["customer_no"])

        # Check EC number groups for multiple identities
        # Use the normalized key for grouping to ensure consistent results (e.g. '0410' == '410')
        ec_groups = matched_records.groupby("_ec_key")
        for ec_key, group in ec_groups:
            if ec_key:
                # Get the original EC number for the message from the first record in the group
                orig_ec = group["EC_NUMBER"].iloc[0]
                
                historical_names = [normalize_name(n) for n in group["CUSTOMER_NAME1"].dropna().unique()]
                all_names_for_ec = list(set(historical_names))
                
                # Only include current application name if it is actually using this EC number
                if app_ec_key == ec_key:
                    if current_name not in all_names_for_ec:
                        all_names_for_ec.append(current_name)
                
                if len(all_names_for_ec) > 1:
                    # First check if names are similar (potential fraud)
                    if detect_similar_names(all_names_for_ec):
                        anomaly_reasons.append(
                            f"Potential identity fraud: EC number '{orig_ec}' used by similar names (possible typos/variations)."
                        )
                        break  # Only flag once per application
                    else:
                        # Names are completely different (confirmed fraud)
                        anomaly_reasons.append(
                            f"Identity fraud detected: EC number '{orig_ec}' used by {len(all_names_for_ec)} different names."
                        )
                        break  # Only flag once per application

        # Check customer number groups for multiple identities
        if not any("Identity fraud detected" in reason or "Potential identity fraud" in reason for reason in anomaly_reasons):
            customer_groups = matched_records.groupby("_customer_key")
            for customer_key, group in customer_groups:
                if customer_key:
                    # Get the original customer number for the message
                    orig_customer = group["CUSTOMER_NO"].iloc[0]
                    
                    historical_names = [normalize_name(n) for n in group["CUSTOMER_NAME1"].dropna().unique()]
                    all_names_for_customer = list(set(historical_names))
                    
                    # Only include current application name if it is actually using this customer number
                    if app_customer_key == customer_key:
                        if current_name not in all_names_for_customer:
                            all_names_for_customer.append(current_name)
                    
                    if len(all_names_for_customer) > 1:
                        # First check if names are similar (potential fraud)
                        if detect_similar_names(all_names_for_customer):
                            anomaly_reasons.append(
                                f"Potential identity fraud: Account number '{orig_customer}' used by similar names (possible typos/variations)."
                            )
                            break  # Only flag once per application
                        else:
                            # Names are completely different (confirmed fraud)
                            anomaly_reasons.append(
                                f"Identity fraud detected: Account number '{orig_customer}' used by {len(all_names_for_customer)} different names."
                            )
                            break  # Only flag once per application

    is_anomaly = len(anomaly_reasons) > 0

    application.update({
        "category": "anomaly" if is_anomaly else "clear",
        "anomaly_reasons": anomaly_reasons,
        "reason": " | ".join(anomaly_reasons) if anomaly_reasons else (
            f"History exists ({int(len(matched_records))} records), no recent conflicts."
            if not matched_records.empty
            else "No matching history found."
        ),
        "reference_date": reference_date.isoformat(),
        "history_match_count": int(len(matched_records)),
        "recent_match_count": int(len(recent_matches)),
        "latest_book_date": latest_book_date_iso,
        "matched_records": to_jsonable_records(
            recent_matches.head(5) if not recent_matches.empty else matched_records.head(5)
        ),
    })
    return application

@app.get("/history/search")
def search_history(query: str):
    query = query.strip()
    if not query:
        return []

    df, _history_warning = load_history_with_status()
    if df.empty:
        return []
    
    # Filter by EC_NUMBER or CUSTOMER_NO or CUSTOMER_NAME1
    # Handle NaNs and convert to string
    df = df.fillna("")
    searchable_columns = ["EC_NUMBER", "CUSTOMER_NO", "CUSTOMER_NAME1"]
    available_columns = [column for column in searchable_columns if column in df.columns]

    if not available_columns:
        return []

    mask = False
    for column in available_columns:
        mask = mask | df[column].astype(str).str.contains(query, case=False, regex=False)

    results = df[mask]
    
    return to_jsonable_records(results)

def build_csv_response_record(row: pd.Series, index: int) -> dict:
    row_number = index + 2

    beneficiary_name = get_first_present(
        row, "BeneficiaryName", "BENEFICIARY NAME", "NAME", "APPLICANT NAME"
    )
    reference = get_first_present(row, "Reference", "REFERENCE", "REF", "ID")
    amount = get_first_present(
        row, "Amount", "AMOUNT", "AMOUNT_FINANCED", "LOAN AMOUNT"
    )
    
    response_status = get_first_present(
        row,
        "BeneficiaryStatus",
        "Status",
        "STATUS",
        "DECISION",
        "RESULT",
        "APPLICATION_STATUS",
    )
    
    initial_status = determine_initial_decision_status(row)

    return {
        "application_id": f"row-{row_number}",
        "row": row_number,
        "applicant_name": beneficiary_name,
        "ec_number": reference,
        "customer_no": reference,
        "amount": normalize_scalar(amount),
        "application_book_date": None,
        "decision_status": initial_status or "pending",
        "response_status": normalize_scalar(response_status),
        "category": "clear",
        "reason": (
            f"CSV response: {normalize_scalar(response_status)}."
            if initial_status is not None
            else "CSV response pending review."
        ),
        "reference_date": datetime.now().isoformat(),
        "history_match_count": 0,
        "recent_match_count": 0,
        "latest_book_date": None,
        "matched_records": [],
    }


def _read_activity_records_unlocked():
    if not ACTIVITY_FILE.exists():
        return []

    try:
        with ACTIVITY_FILE.open("r", encoding="utf-8") as handle:
            return [
                json.loads(line)
                for line in handle
                if line.strip()
            ]
    except (json.JSONDecodeError, OSError):
        return []


def _write_activity_records_unlocked(records: list[dict]):
    ACTIVITY_FILE.parent.mkdir(parents=True, exist_ok=True)
    temp_path = None

    try:
        with NamedTemporaryFile(
            delete=False,
            dir=ACTIVITY_FILE.parent,
            suffix=ACTIVITY_FILE.suffix,
            mode="w",
            encoding="utf-8",
        ) as handle:
            temp_path = Path(handle.name)
            for record in records:
                handle.write(json.dumps(record, ensure_ascii=True))
                handle.write("\n")

        os.replace(temp_path, ACTIVITY_FILE)
    finally:
        if temp_path is not None and temp_path.exists():
            temp_path.unlink(missing_ok=True)


def get_all_activities():
    with ACTIVITY_LOCK:
        return _read_activity_records_unlocked()


def write_activity_records(records: list[dict]):
    with ACTIVITY_LOCK:
        _write_activity_records_unlocked(records)


def get_activity_record_label(record: dict):
    return (
        record.get("applicant_name")
        or record.get("ec_number")
        or record.get("customer_no")
        or record.get("application_id")
    )


def build_record_activity_event(
    record: dict,
    event_type: str,
    timestamp: str,
    *,
    from_status: str | None = None,
    to_status: str | None = None,
    message: str | None = None,
):
    target_status = to_status or record.get("decision_status")
    record_label = get_activity_record_label(record)

    if message is None:
        if event_type == "record_ingested":
            message = f"Record ingested with initial status '{target_status}'."
        elif event_type == "record_decision_updated" and from_status is not None and target_status is not None:
            message = f"Decision changed from '{from_status}' to '{target_status}'."
        else:
            message = event_type.replace("_", " ").title()

    return make_json_safe({
        "event_id": uuid4().hex,
        "timestamp": timestamp,
        "type": event_type,
        "application_id": record.get("application_id"),
        "record_label": record_label,
        "from_status": from_status,
        "to_status": target_status,
        "message": message,
        "reason": record.get("reason"),
        "response_status": record.get("response_status"),
    })


def build_session_activity_event(
    session: dict,
    event_type: str,
    timestamp: str,
    message: str,
    extra: dict | None = None,
):
    payload = {
        "event_id": uuid4().hex,
        "timestamp": timestamp,
        "type": event_type,
        "session_id": session["session_id"],
        "analysis_mode": session["analysis_mode"],
        "file_name": session["file_name"],
        "message": message,
    }
    if extra:
        payload.update(extra)
    return make_json_safe(payload)


def add_record_activity_event(
    session: dict,
    application_id: str,
    event_type: str,
    *,
    timestamp: str | None = None,
    from_status: str | None = None,
    to_status: str | None = None,
    message: str | None = None,
):
    event_timestamp = timestamp or datetime.now().isoformat()
    session["updated_at"] = event_timestamp
    session.setdefault("record_events_by_id", {}).setdefault(application_id, []).append(
        build_record_activity_event(
            session["records_by_id"][application_id],
            event_type,
            event_timestamp,
            from_status=from_status,
            to_status=to_status,
            message=message,
        )
    )
    return event_timestamp


def add_session_activity_event(
    session: dict,
    event_type: str,
    message: str,
    *,
    timestamp: str | None = None,
    extra: dict | None = None,
):
    event_timestamp = timestamp or datetime.now().isoformat()
    session["updated_at"] = event_timestamp
    session.setdefault("activity_events", []).append(
        build_session_activity_event(
            session,
            event_type,
            event_timestamp,
            message,
            extra,
        )
    )
    return event_timestamp


def build_activity_record_snapshot(session: dict, application_id: str):
    record = serialize_record(session["records_by_id"][application_id])
    return make_json_safe({
        **record,
        "source_row": session["rows_by_id"].get(application_id, {}),
        "decision_history": session.get("record_events_by_id", {}).get(application_id, []),
    })


def build_activity_session_snapshot(session: dict):
    records = [
        build_activity_record_snapshot(session, application_id)
        for application_id in session["record_order"]
    ]
    decision_summary = build_decision_summary(session)

    return make_json_safe({
        "session_id": session["session_id"],
        "uploaded_at": session.get("created_at"),
        "updated_at": session.get("updated_at") or session.get("created_at"),
        "analysis_mode": session["analysis_mode"],
        "file_name": session["file_name"],
        "columns": session.get("columns", []),
        "total_processed": session["total_processed"],
        "actionable_records": len(records),
        "record_count": len(records),
        **decision_summary,
        "history_warning": session.get("history_warning"),
        "events": session.get("activity_events", []),
        "records": records,
    })


def normalize_activity_snapshot(activity: dict):
    if activity.get("session_id"):
        return activity

    session_id = activity.get("id")
    timestamp = activity.get("timestamp")
    analysis_mode = activity.get("type")
    file_name = activity.get("filename")
    total_processed = activity.get("total_records", 0)
    approved_count = activity.get("approved_count", 0)
    rejected_count = activity.get("rejected_count", 0)
    pending_count = activity.get("pending_count", 0)

    return {
        "session_id": session_id,
        "uploaded_at": timestamp,
        "updated_at": timestamp,
        "analysis_mode": analysis_mode,
        "file_name": file_name,
        "columns": [],
        "total_processed": total_processed,
        "actionable_records": total_processed,
        "record_count": total_processed,
        "approved_count": approved_count,
        "rejected_count": rejected_count,
        "pending_count": pending_count,
        "history_warning": None,
        "events": [
            {
                "event_id": f"{session_id or 'legacy'}-upload",
                "timestamp": timestamp,
                "type": "legacy_upload",
                "session_id": session_id,
                "analysis_mode": analysis_mode,
                "file_name": file_name,
                "message": f"Legacy activity captured for {file_name or 'upload'}.",
                "total_processed": total_processed,
                "approved_count": approved_count,
                "rejected_count": rejected_count,
                "pending_count": pending_count,
            }
        ],
        "records": [],
    }


def coerce_number(value):
    value = normalize_scalar(value)
    if value is None:
        return 0.0

    try:
        number = float(value)
    except (TypeError, ValueError):
        return 0.0

    if math.isnan(number) or math.isinf(number):
        return 0.0

    return number


def get_activity_datetime(value):
    return parse_date(value)


def get_record_latest_activity_at(record: dict, activity: dict):
    timestamps = [
        get_activity_datetime(event.get("timestamp"))
        for event in record.get("decision_history", [])
    ]
    timestamps.extend([
        get_activity_datetime(activity.get("updated_at")),
        get_activity_datetime(activity.get("uploaded_at")),
    ])
    timestamps = [timestamp for timestamp in timestamps if timestamp is not None]
    return max(timestamps) if timestamps else None


def get_record_status_timestamp(record: dict, target_status: str, fallback: datetime | None):
    matching_timestamps = []
    for event in record.get("decision_history", []):
        to_status = normalize_scalar(event.get("to_status"))
        if str(to_status).lower() != target_status:
            continue

        event_timestamp = get_activity_datetime(event.get("timestamp"))
        if event_timestamp is not None:
            matching_timestamps.append(event_timestamp)

    if matching_timestamps:
        return max(matching_timestamps)

    return fallback


def build_activity_session_summary(activity: dict):
    activity = normalize_activity_snapshot(activity)
    return {
        "session_id": activity.get("session_id"),
        "uploaded_at": activity.get("uploaded_at"),
        "updated_at": activity.get("updated_at"),
        "analysis_mode": activity.get("analysis_mode"),
        "file_name": activity.get("file_name"),
        "total_processed": activity.get("total_processed", 0),
        "actionable_records": activity.get("actionable_records", 0),
        "record_count": activity.get("record_count", activity.get("actionable_records", 0)),
        "approved_count": activity.get("approved_count", 0),
        "rejected_count": activity.get("rejected_count", 0),
        "pending_count": activity.get("pending_count", 0),
        "event_count": len(activity.get("events", [])),
        "history_warning": activity.get("history_warning"),
    }


def initialize_session_activity(session: dict):
    timestamp = datetime.now().isoformat()
    session["created_at"] = timestamp
    session["updated_at"] = timestamp
    session["record_events_by_id"] = {
        application_id: [
            build_record_activity_event(
                session["records_by_id"][application_id],
                "record_ingested",
                timestamp,
            )
        ]
        for application_id in session["record_order"]
    }
    session["activity_events"] = [
        build_session_activity_event(
            session,
            "upload_created",
            timestamp,
            f"Uploaded {session['file_name']} for {session['analysis_mode']} analysis.",
            {
                "total_processed": session["total_processed"],
                "actionable_records": len(session["record_order"]),
                **build_decision_summary(session),
            },
        )
    ]


def upsert_activity(session: dict):
    activity = build_activity_session_snapshot(session)

    with ACTIVITY_LOCK:
        records = _read_activity_records_unlocked()
        updated_records = [
            record
            for record in records
            if record.get("session_id") != activity["session_id"]
            and record.get("id") != activity["session_id"]
        ]
        updated_records.insert(0, activity)
        updated_records.sort(
            key=lambda record: record.get("updated_at") or record.get("uploaded_at") or record.get("timestamp") or "",
            reverse=True,
        )
        _write_activity_records_unlocked(updated_records)


@app.get("/activity")
def get_activity_log():
    activities = get_all_activities()
    activities.sort(
        key=lambda record: record.get("updated_at") or record.get("uploaded_at") or record.get("timestamp") or "",
        reverse=True,
    )
    return [build_activity_session_summary(activity) for activity in activities]


@app.get("/activity/{session_id}")
def get_activity_log_detail(session_id: str):
    activities = get_all_activities()
    for activity in activities:
        if activity.get("session_id") == session_id or activity.get("id") == session_id:
            return normalize_activity_snapshot(activity)

    raise HTTPException(status_code=404, detail="Activity log entry not found.")


@app.get("/reports/approval-ledger")
def get_approval_ledger(days: int = RECENT_APPLICATION_WINDOW_DAYS):
    days = max(1, min(days, 365))
    cutoff = datetime.now() - timedelta(days=days)
    activities = [
        normalize_activity_snapshot(activity)
        for activity in get_all_activities()
    ]

    filtered_records = []
    session_ids = set()

    for activity in activities:
        for record in activity.get("records", []):
            latest_activity_at = get_record_latest_activity_at(record, activity)
            if latest_activity_at is None or latest_activity_at < cutoff:
                continue

            session_id = activity.get("session_id")
            if session_id:
                session_ids.add(session_id)

            decision_status = str(record.get("decision_status") or "pending").lower()
            approved_at = (
                get_record_status_timestamp(record, "approved", latest_activity_at)
                if decision_status == "approved"
                else None
            )

            filtered_records.append(make_json_safe({
                "session_id": session_id,
                "file_name": activity.get("file_name"),
                "analysis_mode": activity.get("analysis_mode"),
                "uploaded_at": activity.get("uploaded_at"),
                "updated_at": activity.get("updated_at"),
                "latest_activity_at": latest_activity_at.isoformat(),
                "approved_at": approved_at.isoformat() if approved_at is not None else None,
                **record,
            }))

    filtered_records.sort(
        key=lambda record: record.get("approved_at") or record.get("latest_activity_at") or "",
        reverse=True,
    )

    approved_records = [
        record for record in filtered_records
        if str(record.get("decision_status") or "").lower() == "approved"
    ]
    rejected_count = sum(
        str(record.get("decision_status") or "").lower() == "declined"
        for record in filtered_records
    )
    pending_count = sum(
        str(record.get("decision_status") or "").lower() in {"pending", "manual_review"}
        for record in filtered_records
    )
    approved_amount = sum(coerce_number(record.get("amount")) for record in approved_records)
    average_amount = approved_amount / len(approved_records) if approved_records else 0.0
    flagged_approved_count = sum(record.get("category") == "anomaly" for record in approved_records)
    clear_approved_count = sum(record.get("category") == "clear" for record in approved_records)

    return make_json_safe({
        "window_days": days,
        "generated_at": datetime.now().isoformat(),
        "session_count": len(session_ids),
        "record_count": len(filtered_records),
        "approved_count": len(approved_records),
        "rejected_count": int(rejected_count),
        "pending_count": int(pending_count),
        "approved_amount": approved_amount,
        "average_amount": average_amount,
        "flagged_approved_count": int(flagged_approved_count),
        "clear_approved_count": int(clear_approved_count),
        "records": approved_records,
    })

@app.post("/upload/approve")
async def approve_funding(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Invalid file format. Please upload an Excel file.")

    contents = await file.read()
    # Read Excel file, skip first row and use second row as headers
    new_df = pd.read_excel(io.BytesIO(contents), header=1)
    history_df, history_lookup, history_warning = load_history_for_matching_with_status()
    session_id = uuid4().hex
    original_columns = list(new_df.columns)
    columns = [str(column) for column in original_columns]
    records_by_id = {}
    rows_by_id = {}
    record_order = []

    for index, row in new_df.iterrows():
        row_values = [normalize_scalar(value) for value in row.tolist()]
        if all(value is None for value in row_values):
            continue

        analyzed_record = analyze_application(row, index, history_df, history_lookup)
        application_id = analyzed_record["application_id"]
        record_order.append(application_id)
        records_by_id[application_id] = analyzed_record
        rows_by_id[application_id] = row_to_storage(row, original_columns)

    session = {
        "session_id": session_id,
        "analysis_mode": "excel",
        "file_name": file.filename,
        "created_at": datetime.now().isoformat(),
        "columns": columns,
        "total_processed": len(new_df),
        "history_warning": history_warning,
        "record_order": record_order,
        "records_by_id": records_by_id,
        "rows_by_id": rows_by_id,
    }
    initialize_session_activity(session)
    ANALYSIS_SESSIONS[session_id] = session
    upsert_activity(session)

    return build_session_response(session)


@app.post("/upload/responses")
async def upload_response_csv(file: UploadFile = File(...)):
    if not file.filename.lower().endswith('.csv'):
        raise HTTPException(status_code=400, detail="Invalid file format. Please upload a CSV file.")

    contents = await file.read()
    new_df = pd.read_csv(io.BytesIO(contents), dtype=object)
    session_id = uuid4().hex
    original_columns = list(new_df.columns)
    columns = [str(column) for column in original_columns]
    records_by_id = {}
    rows_by_id = {}
    record_order = []

    for index, row in new_df.iterrows():
        row_values = [normalize_scalar(value) for value in row.tolist()]
        if all(value is None for value in row_values):
            continue

        analyzed_record = build_csv_response_record(row, index)
        application_id = analyzed_record["application_id"]
        record_order.append(application_id)
        records_by_id[application_id] = analyzed_record
        rows_by_id[application_id] = row_to_storage(row, original_columns)

    session = {
        "session_id": session_id,
        "analysis_mode": "csv",
        "file_name": file.filename,
        "created_at": datetime.now().isoformat(),
        "columns": columns,
        "total_processed": len(new_df),
        "history_warning": None,
        "record_order": record_order,
        "records_by_id": records_by_id,
        "rows_by_id": rows_by_id,
    }
    initialize_session_activity(session)
    ANALYSIS_SESSIONS[session_id] = session
    upsert_activity(session)

    return build_session_response(session)


@app.post("/analysis/{session_id}/records/{application_id}/decision")
def update_record_decision(session_id: str, application_id: str, payload: DecisionPayload):
    session = get_session_or_404(session_id)
    record = session["records_by_id"].get(application_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Record not found in analysis session.")

    decision = payload.decision.strip().lower()
    if decision not in {"pending", "approved", "declined", "manual_review"}:
        raise HTTPException(status_code=400, detail="Unsupported decision.")

    current_status = record["decision_status"]
    if decision == current_status:
        return {
            "record": serialize_record(record),
            "summary": build_decision_summary(session),
        }

    is_excel_session = session.get("analysis_mode") == "excel"
    if is_excel_session:
        if current_status == "approved" and decision != "approved":
            raise HTTPException(
                status_code=409,
                detail="Approved records are already written to history and cannot be changed.",
            )

        if current_status != "approved" and decision == "approved":
            append_record_to_history(session["rows_by_id"][application_id])

    record["decision_status"] = decision
    summary = build_decision_summary(session)
    timestamp = datetime.now().isoformat()
    history_synced = is_excel_session and current_status != "approved" and decision == "approved"
    record_label = get_activity_record_label(record)

    add_record_activity_event(
        session,
        application_id,
        "record_decision_updated",
        timestamp=timestamp,
        from_status=current_status,
        to_status=decision,
        message=(
            f"Decision changed from '{current_status}' to '{decision}'. Record synced to history workbook."
            if history_synced
            else None
        ),
    )
    add_session_activity_event(
        session,
        "record_decision_updated",
        f"{record_label} moved from '{current_status}' to '{decision}'.",
        timestamp=timestamp,
        extra={
            "application_id": application_id,
            "record_label": record_label,
            "from_status": current_status,
            "to_status": decision,
            "history_synced": history_synced,
            **summary,
        },
    )
    upsert_activity(session)

    return {
        "record": serialize_record(record),
        "summary": summary,
    }


@app.get("/analysis/{session_id}/download")
def download_analysis_results(session_id: str):
    session = get_session_or_404(session_id)
    if session.get("analysis_mode") == "csv":
        rejected_df = build_download_dataframe(session, "declined")
        if rejected_df.empty:
            raise HTTPException(status_code=400, detail="No rejected records available for download.")

        output = io.StringIO()
        rejected_df.to_csv(output, index=False)
        csv_bytes = io.BytesIO(output.getvalue().encode("utf-8"))

        download_name = f"{Path(session['file_name']).stem}_rejected.csv"
        return StreamingResponse(
            csv_bytes,
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{download_name}"'},
        )

    approved_df = build_download_dataframe(session, "approved")
    if approved_df.empty:
        raise HTTPException(status_code=400, detail="No approved records available for download.")

    output = io.BytesIO()
    approved_df.to_excel(output, index=False)
    output.seek(0)

    download_name = f"{Path(session['file_name']).stem}_approved.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{download_name}"'},
    )


@app.get("/analysis/{session_id}/approved/download")
def download_approved_records(session_id: str):
    return download_analysis_results(session_id)

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
